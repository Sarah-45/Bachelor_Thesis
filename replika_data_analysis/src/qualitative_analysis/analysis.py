from collections import Counter
from itertools import combinations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy import stats

from src.config import (
    ANALYSIS_RESULTS_FILE,
    CLAUDE_RELEVANT_POSTS_FILE,
    LLM_CODED_FILE,
    QUALITATIVE_DIR,
)
from src.constants import BROAD_EMOTIONAL_WORDS, CODEBOOK_MARKERS

CODES = [
    "C1_separation_distress",
    "C2_fear_abandonment",
    "C3_proximity_seeking",
    "C4_reassurance_seeking",
    "C5_romantic_dependency",
]
CODE_NAMES = {
    "C1_separation_distress": "C1 Separation Distress",
    "C2_fear_abandonment":    "C2 Fear of Abandonment",
    "C3_proximity_seeking":   "C3 Proximity Seeking",
    "C4_reassurance_seeking": "C4 Reassurance Seeking",
    "C5_romantic_dependency": "C5 Romantic Dependency",
}



def _style_header(cell) -> None:
    cell.font      = Font(bold=True, color="000000", name="Arial", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )


def _style_cell(cell, shade: bool = False) -> None:
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(row=row, column=col, value=h))


def _write_data_row(ws, row: int, values: list, shade: bool = False) -> None:
    for col, val in enumerate(values, 1):
        _style_cell(ws.cell(row=row, column=col, value=val), shade)



def run_analysis() -> dict:
    print("Loading data...")
    llm      = pd.read_csv(LLM_CODED_FILE)
    original = pd.read_csv(CLAUDE_RELEVANT_POSTS_FILE)

    df    = llm[llm["relevant"] == True].copy()
    total = len(df)
    print(f"Relevant posts: {total}")

    wb = Workbook()

    print("1. Code frequency...")
    ws1 = wb.active
    ws1.title = "1. Code Frequency"
    _write_header_row(ws1, 1, ["Code", "Name", "Count", "% of relevant posts"])
    _set_col_widths(ws1, [8, 30, 10, 20])
    ws1.row_dimensions[1].height = 25

    for i, code in enumerate(CODES, 2):
        count = int(df[code].sum())
        pct   = count / total * 100
        _write_data_row(ws1, i, [code.split("_")[0].upper(), CODE_NAMES[code], count, f"{pct:.1f}%"], shade=i % 2 == 0)

    _write_data_row(ws1, len(CODES) + 2, ["", "Total relevant posts", total, "100%"])
    for col in range(1, 5):
        ws1.cell(row=len(CODES) + 2, column=col).font = Font(bold=True, name="Arial", size=10)

    print("2. Code overlap matrix...")
    ws2       = wb.create_sheet("2. Code Overlap Matrix")
    short_names = ["C1", "C2", "C3", "C4", "C5"]
    ws2.cell(row=1, column=1, value="Code Overlap (posts where both codes appear together)")
    ws2.cell(row=1, column=1).font = Font(bold=True, name="Arial", size=11)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    _write_header_row(ws2, 2, [""] + short_names)

    for i, code_r in enumerate(CODES, 3):
        ws2.cell(row=i, column=1, value=short_names[i - 3])
        _style_header(ws2.cell(row=i, column=1))
        for j, code_c in enumerate(CODES, 2):
            cell = ws2.cell(row=i, column=j)
            if code_r == code_c:
                cell.value = int(df[code_r].sum())
                cell.font  = Font(bold=True, name="Arial", size=10, color="000000")
            else:
                cell.value = int((df[code_r] & df[code_c]).sum())
                _style_cell(cell, shade=(i + j) % 2 == 0)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
            )
    _set_col_widths(ws2, [8, 12, 12, 12, 12, 12])

    print("3. Co-occurrence significance...")
    ws3 = wb.create_sheet("3. Co-occurrence Significance")
    _write_header_row(ws3, 1, ["Code A", "Code B", "Both present", "Expected by chance", "Chi-square", "p-value", "Significant?"])
    _set_col_widths(ws3, [25, 25, 14, 20, 12, 12, 14])
    ws3.row_dimensions[1].height = 25

    row = 2
    for code_a, code_b in combinations(CODES, 2):
        both    = int((df[code_a] & df[code_b]).sum())
        only_a  = int((df[code_a] & ~df[code_b]).sum())
        only_b  = int((~df[code_a] & df[code_b]).sum())
        neither = int((~df[code_a] & ~df[code_b]).sum())

        try:
            chi2, p, _, expected = stats.chi2_contingency([[both, only_a], [only_b, neither]])
            expected_both = round(expected[0][0], 1)
            significant   = "Yes" if p < 0.05 else "No"
            p_str         = f"{p:.4f}"
            chi_str       = f"{chi2:.2f}"
        except Exception:
            expected_both, significant, p_str, chi_str = "-", "-", "-", "-"

        _write_data_row(ws3, row, [
            CODE_NAMES[code_a], CODE_NAMES[code_b], both, expected_both, chi_str, p_str, significant,
        ], shade=row % 2 == 0)

        if significant == "Yes":
            ws3.cell(row=row, column=7).font = Font(bold=True, name="Arial", size=10, color="1B5E20")
        row += 1

    print("4. Codebook marker frequency...")
    ws4 = wb.create_sheet("4. Codebook Markers")
    df_text  = df.merge(original[["id", "reddit_post_text"]], on="id", how="left")
    all_text = " ".join(df_text["reddit_post_text"].fillna("").str.lower().tolist())

    _write_header_row(ws4, 1, ["Code", "Marker", "Count in dataset"])
    _set_col_widths(ws4, [25, 45, 18])
    ws4.row_dimensions[1].height = 25

    row = 2
    for code_name, markers in CODEBOOK_MARKERS.items():
        for marker in markers:
            count = all_text.count(marker.lower())
            _write_data_row(ws4, row, [code_name, marker, count], shade=row % 2 == 0)
            row += 1

    print("5. Broad linguistic analysis...")
    ws5 = wb.create_sheet("5. Broad Linguistic")
    _write_header_row(ws5, 1, ["Word", "Count in dataset", "% of posts containing word"])
    _set_col_widths(ws5, [25, 18, 28])
    ws5.row_dimensions[1].height = 25

    word_counts = []
    for word in BROAD_EMOTIONAL_WORDS:
        count      = all_text.count(word.lower())
        posts_with = df_text["reddit_post_text"].fillna("").str.lower().str.contains(word.lower()).sum()
        pct        = posts_with / total * 100
        word_counts.append((word, count, pct))

    word_counts.sort(key=lambda x: x[1], reverse=True)
    for i, (word, count, pct) in enumerate(word_counts, 2):
        _write_data_row(ws5, i, [word, count, f"{pct:.1f}%"], shade=i % 2 == 0)

    print("6. Dominant combinations...")
    ws6 = wb.create_sheet("6. Dominant Combinations")
    _write_header_row(ws6, 1, ["Code combination", "Count", "% of relevant posts"])
    _set_col_widths(ws6, [45, 10, 22])
    ws6.row_dimensions[1].height = 25

    combos: Counter = Counter()
    for _, row_data in df.iterrows():
        active = tuple(sorted([code.split("_")[0].upper() for code in CODES if row_data[code]]))
        if active:
            combos[active] += 1

    for i, (combo, count) in enumerate(combos.most_common(20), 2):
        combo_str = " + ".join(combo) if combo else "none"
        pct       = count / total * 100
        _write_data_row(ws6, i, [combo_str, count, f"{pct:.1f}%"], shade=i % 2 == 0)

    print("7. Trauma context...")
    ws7        = wb.create_sheet("7. Trauma Context")
    trauma     = df[df["trauma_context"] == True]
    no_trauma  = df[df["trauma_context"] == False]
    _write_header_row(ws7, 1, ["", "All posts", "Trauma context", "No trauma context"])
    _set_col_widths(ws7, [30, 14, 18, 22])
    ws7.row_dimensions[1].height = 25
    _write_data_row(ws7, 2, ["Total posts", total, len(trauma), len(no_trauma)])
    ws7.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for i, code in enumerate(CODES, 3):
        all_pct = df[code].sum() / total * 100
        t_pct   = trauma[code].sum() / len(trauma) * 100    if len(trauma) > 0    else 0
        nt_pct  = no_trauma[code].sum() / len(no_trauma) * 100 if len(no_trauma) > 0 else 0
        _write_data_row(ws7, i, [
            CODE_NAMES[code],
            f"{int(df[code].sum())} ({all_pct:.1f}%)",
            f"{int(trauma[code].sum())} ({t_pct:.1f}%)",
            f"{int(no_trauma[code].sum())} ({nt_pct:.1f}%)",
        ], shade=i % 2 == 0)
        ws7.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center")

    print("8. High vs Medium confidence...")
    ws8 = wb.create_sheet("8. High vs Medium Confidence")
    original["id"] = original["id"].astype(str)
    df["id"]       = df["id"].astype(str)
    merged_conf    = df.merge(original[["id", "confidence"]], on="id", how="left")

    high   = merged_conf[merged_conf["confidence"] == "High"]
    medium = merged_conf[merged_conf["confidence"] == "Medium"]
    _write_header_row(ws8, 1, ["", "High confidence", "Medium confidence"])
    _set_col_widths(ws8, [30, 20, 22])
    ws8.row_dimensions[1].height = 25
    _write_data_row(ws8, 2, ["Total posts", len(high), len(medium)])
    ws8.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

    avg_high   = high[CODES].sum(axis=1).mean()   if len(high) > 0   else 0
    avg_medium = medium[CODES].sum(axis=1).mean() if len(medium) > 0 else 0
    _write_data_row(ws8, 3, ["Avg codes per post", f"{avg_high:.2f}", f"{avg_medium:.2f}"], shade=True)
    ws8.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")

    for i, code in enumerate(CODES, 4):
        h_pct = high[code].sum() / len(high) * 100     if len(high) > 0   else 0
        m_pct = medium[code].sum() / len(medium) * 100 if len(medium) > 0 else 0
        _write_data_row(ws8, i, [
            CODE_NAMES[code],
            f"{int(high[code].sum())} ({h_pct:.1f}%)",
            f"{int(medium[code].sum())} ({m_pct:.1f}%)",
        ], shade=i % 2 == 0)
        ws8.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ANALYSIS_RESULTS_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ANALYSIS_RESULTS_FILE)
    print(f"\nDONE — saved to {ANALYSIS_RESULTS_FILE}")

    return {"relevant_posts_analysed": total, "output": str(ANALYSIS_RESULTS_FILE)}


if __name__ == "__main__":
    run_analysis()
